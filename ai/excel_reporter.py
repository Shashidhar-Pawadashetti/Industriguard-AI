import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from datetime import datetime
import os

class ExcelReporter:
    def __init__(self, report_path="reports/employee_safety.xlsx"):
        self.report_path = report_path
        os.makedirs("reports", exist_ok=True)

        # Load existing or create new
        if os.path.exists(report_path):
            self.wb = openpyxl.load_workbook(report_path)
            self.ws = self.wb.active
            print(f"[ExcelReporter] Loaded existing report → {report_path}")
        else:
            self._create_new_report()
            print(f"[ExcelReporter] Created new report → {report_path}")

    def _create_new_report(self):
        """Creates a fresh Excel file with headers"""
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "Safety Report"

        # ── Column widths ─────────────────────────────────────────
        self.ws.column_dimensions["A"].width = 12  # Employee ID
        self.ws.column_dimensions["B"].width = 25  # Name
        self.ws.column_dimensions["C"].width = 20  # Department
        self.ws.column_dimensions["D"].width = 12  # Helmet
        self.ws.column_dimensions["E"].width = 14  # Safety Vest
        self.ws.column_dimensions["F"].width = 14  # Status
        self.ws.column_dimensions["G"].width = 22  # Timestamp
        self.ws.column_dimensions["H"].width = 25  # Notes

        # ── Title row ─────────────────────────────────────────────
        self.ws.merge_cells("A1:H1")
        title_cell = self.ws["A1"]
        title_cell.value           = "IndustriGuard AI — Employee Safety Report"
        title_cell.font            = Font(
            name="Calibri", size=14,
            bold=True, color="FFFFFF"
        )
        title_cell.fill            = PatternFill(
            fill_type="solid", fgColor="0D3B6E"
        )
        title_cell.alignment       = Alignment(
            horizontal="center", vertical="center"
        )
        self.ws.row_dimensions[1].height = 35

        # ── Date row ──────────────────────────────────────────────
        self.ws.merge_cells("A2:H2")
        date_cell = self.ws["A2"]
        date_cell.value     = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        date_cell.font      = Font(name="Calibri", size=10, italic=True)
        date_cell.alignment = Alignment(horizontal="center")

        # ── Header row ────────────────────────────────────────────
        headers = [
            "Employee ID", "Name", "Department",
            "Helmet", "Safety Vest", "Status",
            "Timestamp", "Notes"
        ]
        header_fill = PatternFill(fill_type="solid", fgColor="1A5276")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        border      = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        for col_num, header in enumerate(headers, 1):
            cell            = self.ws.cell(row=3, column=col_num, value=header)
            cell.font       = header_font
            cell.fill       = header_fill
            cell.alignment  = Alignment(horizontal="center", vertical="center")
            cell.border     = border

        self.ws.row_dimensions[3].height = 25
        self._save()

    def _find_employee_row(self, employee_id):
        """Searches for existing row with this employee ID"""
        for row in self.ws.iter_rows(min_row=4):
            if row[0].value == employee_id:
                return row[0].row
        return None

    def _get_next_empty_row(self):
        """Returns the next empty row after all data"""
        return self.ws.max_row + 1

    def update_employee(self, employee, status_data):
        """
        Updates or adds a row for this employee.
        If employee already has a row → update it.
        If new employee → add new row.
        """
        emp_id      = employee["id"]
        emp_name    = employee["name"]
        emp_dept    = employee["department"]
        has_helmet  = "✓" if status_data["has_helmet"] else "✗"
        has_vest    = "✓" if status_data["has_vest"] else "✗"
        status      = status_data["status"]
        timestamp   = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        notes       = status_data["message"]

        # Find existing row or get new row number
        row_num = self._find_employee_row(emp_id)
        if not row_num:
            row_num = self._get_next_empty_row()
            print(f"[ExcelReporter] Adding new row for {emp_id}")
        else:
            print(f"[ExcelReporter] Updating row {row_num} for {emp_id}")

        # ── Row data ──────────────────────────────────────────────
        row_data = [
            emp_id, emp_name, emp_dept,
            has_helmet, has_vest, status,
            timestamp, notes
        ]

        # ── Row styling ───────────────────────────────────────────
        ready_fill     = PatternFill(fill_type="solid", fgColor="D5F5E3")
        not_ready_fill = PatternFill(fill_type="solid", fgColor="FADBD8")
        row_fill       = ready_fill if status == "READY" else not_ready_fill

        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        for col_num, value in enumerate(row_data, 1):
            cell            = self.ws.cell(row=row_num, column=col_num, value=value)
            cell.fill       = row_fill
            cell.border     = border
            cell.alignment  = Alignment(horizontal="center", vertical="center")
            cell.font       = Font(name="Calibri", size=11)

            # Special color for status cell
            if col_num == 6:
                if status == "READY":
                    cell.font = Font(
                        name="Calibri", size=11,
                        bold=True, color="1E8449"
                    )
                else:
                    cell.font = Font(
                        name="Calibri", size=11,
                        bold=True, color="C0392B"
                    )

            # Red color for missing PPE cells
            if col_num in [4, 5] and value == "✗":
                cell.font = Font(
                    name="Calibri", size=13,
                    bold=True, color="C0392B"
                )
            elif col_num in [4, 5] and value == "✓":
                cell.font = Font(
                    name="Calibri", size=13,
                    bold=True, color="1E8449"
                )

        self.ws.row_dimensions[row_num].height = 22
        self._save()

        print(f"[ExcelReporter] Saved → {emp_id} | {status}")

    def _save(self):
        self.wb.save(self.report_path)